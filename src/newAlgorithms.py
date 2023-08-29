import os
# import urllib
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import magic


def getFiles(dir_path):
    listFiles = []
    try:
        for path, subdirs, files in os.walk(dir_path):
            for name in files:
                filePath = os.path.join(path, name)
                relativePathDir = filePath.replace(dir_path,"")
                pathWithoutFilename = relativePathDir.replace(name,"")
                # hàm lấy giá trị filesize (đơn vị byte)
                fileSize = os.stat(filePath).st_size
                fileType = detectHeaderFile(filePath)
                headerfile = getHeaderFile(filePath)
                element = [pathWithoutFilename, name, relativePathDir, fileSize, headerfile, fileType]
                # element = [pathWithoutFilename, name, relativePathDir, fileSize, fileType]
                # print(element)
                listFiles.append(element)
    except FileNotFoundError:
        print(f"The directory {dir_path} does not exist")
    except PermissionError:
        print(f"Permission denied to access the directory {dir_path}")
    except OSError as e:
        print(f"An OS error occurred: {e}")
    return listFiles
# Note: Chưa xử lý đc chỗ điều kiện exception

def compare2dir(list1, list2):
    result = []
    serial = 0
    cloneList1 = list1.copy()
    cloneList2 = list2.copy()
    for i in cloneList2:
        for j in cloneList1:
            if i[2] == j[2]:
                # So sánh file trước và file sau
                changedSize = i[3]-j[3]
                # Nếu dung lượng không thay đổi -> bỏ qua
                if changedSize == 0:
                    cloneList1.remove(j)
                    break
                # Nếu dung lượng file thay đổi -> lưu bản ghi
                else:
                    serial += 1
                    # Đổi đơn vị từ byte thành KB
                    newSizeKB = round (j[3] / 1024, 1) 
                    oldSizeKB = round (i[3] / 1024, 1) 
                    element =  [serial, j[2],"Sửa", newSizeKB, oldSizeKB, changedSize, "Thay đổi thông tin", i[4], i[5]]
                    result.append(element)
                    cloneList1.remove(j)
                    break
            else:
                if j == cloneList1[-1]:
                    serial += 1
                    filesizeKB = round (i[3] / 1024, 1) 
                    element =  [serial, i[2],"Thêm mới", None, filesizeKB, i[3], "Tạo mới", i[4], i[5]]
                    result.append(element)
                    break
                else:
                    continue
    for k in cloneList1:
        serial += 1
        filesizeKB = round (k[3] / 1024, 1)
        element =  [serial, k[2],"Xóa", filesizeKB, None, -k[3], "Không còn sử dụng", i[4], i[5]]
        result.append(element)

    return result

def countFiles(listFiles):
    return str(len(listFiles))

def totalSize(listFiles):
    total = 0
    for i in listFiles:
        total += i[3]
    totalMB = round(total / (1024*1024), 4)
    return str(totalMB)

def writeToExcelFile(filesDir1, filesDir2, resultList, dir1, dir2, version, productName, outputFile):
    wb = openpyxl.Workbook()
    # wb = openpyxl.load_workbook(outputFile)
    ws = wb.active
    ws["A1"] = "BẢNG KHAI BÁO CẤU TRÚC VÀ DUNG LƯỢNG PHÁT HÀNH SẢN PHẨM: " + productName
    ws.merge_cells("A1:I1")
    ws["A2"] = "Phiên bản:"
    ws["B2"] = dir1
    ws["F2"] = "Số files: " +  countFiles(filesDir1) + " - Dung lượng: " + totalSize(filesDir1) + "MB"
    ws["B3"] = dir2
    ws["F3"] = "Số files: " +  countFiles(filesDir2) + " - Dung lượng: " + totalSize(filesDir2) + "MB"
    fieldNames = ["STT", version, "Tình trạng", "Dung lượng cũ (KB)", "Dung lượng mới (KB)", "Chênh lệch (B)", "Mục đích sử dụng", "Header file", "Kiểu file"]
    # fieldNames = ["STT", version, "Tình trạng", "Dung lượng cũ (KB)", "Dung lượng mới (KB)", "Chênh lệch (B)", "Mục đích sử dụng", "Kiểu file"]
    ws.append(fieldNames)
    for i in resultList:
        ws.append(i)

    # Create a few styles
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    big_bold_blue_font = Font(color = '173B9E', bold=True, size = 14)
    double_border_side = Side(border_style='thin')
    square_border = Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
    bold12 = Font(bold=True, size = 12)
    gay_background = PatternFill('solid', start_color = 'B1B1B2')

    ws["A1"].font = big_bold_blue_font
    ws["A1"].alignment = center_aligned_text
    ws["A2"].font = bold12
    ws.row_dimensions[4].height = 28

    for cell in ws[4]:
        cell.font = bold12
        cell.fill = gay_background
        cell.border = square_border
        cell.alignment = center_aligned_text

    for i in range (4, ws.max_row+1):
        ws["A"+ str(i)].alignment = center_aligned_text

    # Custom column width size
    ws.column_dimensions["A"].width = len(ws["A2"].value) + 2


    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', "I"]
    for index, column_letter in enumerate(columns):
        values = [ws.cell(row=i,column=index+2).value for i in range(1,ws.max_row+1)]
        max_width = 0
        for value in values:
            if value == None: continue
            elif max_width <= len(str(value)): max_width = len(value) + 4
            # elif max_width <= len(str(value)): max_width = len(value) - 4
        ws.column_dimensions[str(column_letter)].width = max_width
    
    wb.save(filename=outputFile)

def detectHeaderFile(pathToFile):
    
    # fileType = magic.from_file(pathToFile)
    # fileType = magic.from_file(pathToFile)
    fileType = magic.from_buffer(open(pathToFile, "rb").read(2048))
    return fileType

def getHeaderFile(pathToFile):
    with open(pathToFile, "rb") as f:
        header = f.read(2048)
    return str(header)


# def readExcelFile(filePath):
#     wb = openpyxl.load_workbook(filePath)
#     ws = wb.active
#     max_row = ws.max_row
#     max_col = ws.max_column
    
#     listRecord = []
#     for i in range(5, max_row + 1):
#         record = []
#         for j in range(2, max_col+1):
#             cell_val = ws.cell(row = i, column = j).value
#             record.append(cell_val)
#         listRecord.append(record)
#     return listRecord


# def compare2list(list1, list2):
# # list1 is result of 'compare2dir' function
# # list2 is result of 'readExcelFile' function
#     gateBlocking = False
#     cloneList1 = list1.copy()
#     cloneList2 = list2.copy()
#     for i in cloneList1:
#         for j in cloneList2:
#             if i[0] == j[0]:
#                 if i[-2] == j[-2]:
#                     i.append("pass")
#                     i.append(None)
#                 else:
#                     i.append("fail")
#                     i.append("Dung lượng file thay đổi khác với khai báo")
#                     gateBlocking = True
#                 cloneList2.remove(j)

#             elif j == cloneList2[-1] and i[0] != j[0]:
#                 temp = j.copy()
#                 temp.append("fail")
#                 temp.append("File mới nằm ngoài danh sách khai báo")
#                 cloneList1.append(temp)
#                 cloneList2.remove(j)
#                 gateBlocking = True
#     if cloneList2 != []:
#         for k in cloneList2:
#             k.append("fail")
#             k.append("File tồn tại nhưng không được khai báo")
#             cloneList1.append(k)
#             gateBlocking = True

#     return cloneList1